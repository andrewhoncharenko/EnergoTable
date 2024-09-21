# -*- coding: utf-8 -*-
import psycopg2
import openpyxl
import sys
import os
import array as a
from copy import copy
from datetime import datetime
from datetime import timedelta
from calendar import monthrange
from openpyxl.styles import Font
from openpyxl.styles import PatternFill, Border, Alignment

def main():
    arguments = [];
    argumentValues = [];
    host = ""
    port = ""
    user = ""
    password = ""
    database = ""
    filePath = ""
    month = 1
    year = 1
    path = ""
    index = 1
    for argument in sys.argv:
        if argument.startswith("-"):
            arguments.append(argument)
        else:
            argumentValues.append(argument)
    for argument in arguments:
        if argument == "-h":
            host = argumentValues[index]
        if argument == "-p":
            port = argumentValues[index]
        if argument == "-u":
            user = argumentValues[index]
        if argument == "-pw":
            password = argumentValues[index]
        if argument == "-db":
            database = argumentValues[index]
        if argument == "-f":
            filePath = argumentValues[index]
        if argument == "-m":
            month = argumentValues[index]
        if argument == "-y":
            year = argumentValues[index]
        index += 1

    path = os.path.dirname(__file__)
    
    connection = psycopg2.connect(host = host, port = port, user = user, password = password, dbname = database)

    export_to_excel(connection, filePath, path, month, year)

    connection.close()

def fillFeeders(databaseConnection, sheet, feeders, block1Length, startColumn, startDate, endDate, font, fillColor):
    offset = 0
    feederIdsString = ""
    feederBlockLength = 52
    for feeder in enumerate(feeders, start = 2):
        feederRow = block1Length + offset
        feederId = feeder[1][0]
        queryString = "select real_val, dt from d_gr30 where dt >= '" + startDate.strftime("%Y-%m-%d %H:%M:%S") + "' and dt <= '" + endDate.strftime("%Y-%m-%d %H:%M:%S")
        queryString += "' and idfiders = " + str(feederId) + " and dr = 1 ORDER BY dt"
        dataCursor = databaseConnection.cursor()
        dataCursor.execute(queryString)
        gr30 = dataCursor.fetchall()
        dataCursor.close()
        pDay = 0
        column = startColumn
        daySum = 0
        feederName = feeder[1][1]
        print(feederName)
        feederIdsString += str(feederId) + ","
        workCell = sheet.cell(feederRow, column)
        workCell.font = font
        workCell.value = feederName
        feederRow += 1
        workCell = sheet.cell(feederRow, column)
        workCell.fill = fillColor
        workCell.value = "Время"
        feederRow += 1
        column += 1
        for gr30value in enumerate(gr30, start = 2):
            if pDay != gr30value[1][1].day:
                if pDay != 0:
                    workCell = sheet.cell(feederRow, column)
                    workCell.value = daySum
                    if column == startColumn + 1:
                        workCell = sheet.cell(feederRow, column - 1)
                        workCell.value = "Всего"
                    daySum = 0
                    feederRow = block1Length + offset + 2
                    column += 1
                workCell = sheet.cell(feederRow - 1, column)
                workCell.fill = fillColor
                workCell.value = gr30value[1][1].strftime("%d") + "." + gr30value[1][1].strftime("%m") + "." + gr30value[1][1].strftime("%Y")[2:]
            if column == startColumn + 1:
                workCell = sheet.cell(feederRow, column - 1)
                workCell.value = (gr30value[1][1] + timedelta(minutes = 30)).strftime("%H:%M")
            workCell = sheet.cell(feederRow, column)
            value = gr30value[1][0]
            if value == None:
                value = 0
            workCell.value = value
            daySum += value
            feederRow += 1
            pDay = gr30value[1][1].day
        workCell = sheet.cell(block1Length + offset + 1, column + 1)
        workCell.fill = fillColor
        workCell.value = "Всего"
        workCell = sheet.cell(feederRow, column)
        workCell.value = daySum
        offset += feederBlockLength
    return feederIdsString

def fillSvod(sheet, title, objects, cRow, fill):
    sheet.title = title
    sheetRow = 3 
    baseSheetRow = 3
    sheetColumn = 2
    maxColumn = 33
    cr = cRow
    pRow = sheetRow
    workCell = sheet.cell(sheetRow - 2, sheetColumn - 1)
    workCell.value = "СВОДНАЯ ПО " + title + " СОБСТВЕННОЕ"
    coordinateCell = sheet.cell(baseSheetRow, sheetColumn - 1)
    workCell = sheet.cell(sheetRow, sheetColumn - 1)
    workCell.value = "='" + objects[0] + "'!" + coordinateCell.coordinate
    for row in range(sheetRow, 53):
        for column in range(sheetColumn, maxColumn):
            if baseSheetRow == 3:
                coordinateCell = sheet.cell(baseSheetRow, column)
                workCell = sheet.cell(row, column)
                workCell.fill = fill
                workCell.value = "='" + objects[0] + "'!" + coordinateCell.coordinate
            else:
                if column == 2:
                    coordinateCell = sheet.cell(baseSheetRow, column - 1)
                    workCell = sheet.cell(row, column - 1)
                    workCell.value = "='" + objects[0] + "'!" + coordinateCell.coordinate
                workCell = sheet.cell(row, column)
                
                value = "="
                for sObject in objects:
                    if sObject == "Чумаченко, 13В":
                        if pRow != row:
                            cr += 1
                            pRow = row
                        coordinateCell = sheet.cell(cr, column)
                        cellCoordinate = coordinateCell.coordinate
                    else:
                        coordinateCell = sheet.cell(baseSheetRow, column)
                        cellCoordinate = coordinateCell.coordinate
                    value += "'" + sObject + "'!" + cellCoordinate + "+"
                workCell.value = value[:len(value) - 1] 
        baseSheetRow += 1
    sheetRow = 56
    baseSheetRow = 56
    pRow = sheetRow
    cr = cRow
    workCell = sheet.cell(sheetRow - 2, sheetColumn - 1)
    workCell.value = "СВОДНАЯ ПО " + title
    for row in range(sheetRow, 103):
        for column in range(sheetColumn, maxColumn):
            if baseSheetRow == 3:
                coordinateCell = sheet.cell(baseSheetRow, column)
                workCell = sheet.cell(row, column)
                workCell.value = "='" + objects[0] + "'!" + coordinateCell.coordinate
            else:
                if column == 2:
                    coordinateCell = sheet.cell(baseSheetRow, column - 1)
                    workCell = sheet.cell(row, column - 1)
                    workCell.value = "='" + objects[0] + "'!" + coordinateCell.coordinate
                workCell = sheet.cell(row, column)
                value = "="
                for sObject in objects:
                    if sObject == "Чумаченко, 13В":
                        if pRow != row:
                            cr += 1
                            pRow = row
                        coordinateCell = sheet.cell(cr, column)
                        cellCoordinate = coordinateCell.coordinate
                    else:
                        coordinateCell = sheet.cell(baseSheetRow, column)
                        cellCoordinate = coordinateCell.coordinate
                    value += "'" + sObject + "'!" + cellCoordinate + "+"
                workCell.value = value[:len(value) - 1]
        baseSheetRow += 1

def fillSubSvod(sheet, title, objects, coordinateSheet):
    sheet.title = title
    maxColumn = 33
    maxRow = 52
    sheetRow = 3
    sheetColumn = 1
    baseSheetRow = 56
    baseSheetColumn = 35
    workCell = sheet.cell(sheetRow - 2, sheetColumn)
    workCell.value = "СВОДНАЯ ПО СУБПОТРЕБИТЕЛЯМ"
    for row in range(sheetRow, maxRow):
        for column in range(sheetColumn, maxColumn + 1):
            if baseSheetRow == 3:
                coordinateCell = sheet.cell(baseSheetRow, column)
                workCell = sheet.cell(row, column)
                workCell.value = "='" + objects[0] + "'!" + coordinateCell.coordinate
            else:
                if column == 2:
                    coordinateCell = sheet.cell(baseSheetRow, column - 1)
                    workCell = sheet.cell(row, column - 1)
                    workCell.value = "='" + objects[0] + "'!" + coordinateCell.coordinate
                workCell = sheet.cell(row, column)
                coordinateCell = coordinateSheet.cell(baseSheetRow, baseSheetColumn)
                cellCoordinate = coordinateCell.coordinate
                value = "="
                for sObject in objects:
                    value += "'" + sObject + "'!" + cellCoordinate + "+"
                workCell.value = value[:len(value) - 1]
            baseSheetColumn += 1
        baseSheetColumn = 35
        baseSheetRow += 1

def fillNoASKUE(connection, book, startDate, endDate, font, fill):
    objects = [88, 90, 365]
    objectNames = ["Суб без АСКУЭ договор 10", "Суб без АСКУЭ договор 15", "Суб без АСКУЭ договор 2"]
    i = 0 
    for objectId in objects:
        queryString = "select dt, val from d_gr30 where dt >= '" + startDate.strftime("%Y-%m-%d %H:%M:%S") + "' and dt <= '" + endDate.strftime("%Y-%m-%d %H:%M:%S") + "' and idfiders = " + str(objectId)
        dataCursor = connection.cursor()
        dataCursor.execute(queryString)
        gr30 = dataCursor.fetchall()
        dataCursor.close()
        sheetRow = 1
        sheetColumn = 1
        sheet = book.create_sheet(objectNames[i])
        workCell = sheet.cell(sheetRow, sheetColumn)
        workCell.font = font
        workCell.value = objectNames[i]
        sheetRow += 2
        workCell = sheet.cell(sheetRow, sheetColumn)
        workCell.fill = fill
        workCell.value = "Время"
        pDay = 0
        daySum = 0
        for data in enumerate(gr30, start = 2):
            if pDay != data[1][0].day:
                if pDay != 0:
                    workCell = sheet.cell(sheetRow, sheetColumn)
                    workCell.value = daySum
                sheetRow = 3
                sheetColumn += 1
                workCell = sheet.cell(sheetRow, sheetColumn)
                workCell.fill = fill
                workCell.value = data[1][0].strftime("%d") + "." + data[1][0].strftime("%m") + "." + data[1][0].strftime("%Y")[2:]
                sheetRow += 1
                daySum = 0
            value = data[1][1]
            if value == None:
                value = 0
            workCell = sheet.cell(sheetRow, sheetColumn)
            workCell.value = value
            daySum += value
            pDay = data[1][0].day
            sheetRow += 1
        workCell = sheet.cell(sheetRow, 1)
        workCell.value = "Всего"
        workCell = sheet.cell(sheetRow, sheetColumn)
        workCell.value = daySum
        i += 1
    
def fillSvodSheet(connection, book, startDate, endDate, font, fill):
    queryString = "select dt, val from d_gr30 where dt >= '" + startDate.strftime("%Y-%m-%d %H:%M:%S") + "' and dt <= '" + endDate.strftime("%Y-%m-%d %H:%M:%S") + "' and idfiders = 418"
    dataCursor = connection.cursor()
    dataCursor.execute(queryString)
    gr30 = dataCursor.fetchall()
    dataCursor.close()
    sheetRow = 1
    sheetColumn = 1
    sheet = book.create_sheet("Сводная")
    workCell = sheet.cell(sheetRow, sheetColumn)
    workCell.font = font
    workCell.value = "Сводная"
    sheetRow += 2
    workCell = sheet.cell(sheetRow, sheetColumn)
    workCell.fill = fill
    workCell.value = "Время"
    pDay = 0
    daySum = 0
    for data in enumerate(gr30, start = 2):
        if pDay != data[1][0].day:
            if pDay != 0:
                workCell = sheet.cell(sheetRow, sheetColumn)
                workCell.value = daySum
            sheetRow = 3
            sheetColumn += 1
            workCell = sheet.cell(sheetRow, sheetColumn)
            workCell.fill = fill
            workCell.value = data[1][0].strftime("%d") + "." + data[1][0].strftime("%m") + "." + data[1][0].strftime("%Y")[2:]
            sheetRow += 1
            daySum = 0
        value = data[1][1]
        if value == None:
            value = 0
        workCell = sheet.cell(sheetRow, sheetColumn)
        workCell.value = value
        daySum += value
        pDay = data[1][0].day
        sheetRow += 1
    workCell = sheet.cell(sheetRow, 1)
    workCell.value = "Всего"
    workCell = sheet.cell(sheetRow, sheetColumn)
    workCell.value = daySum

def fillC1(sheet, objects, templateSheet):
    maxColumn = templateSheet.max_column
    maxRow = templateSheet.max_row
    for row in range(1, maxRow + 1):
        for column in range(1, maxColumn + 1):
            templateCell = templateSheet.cell(row, column)
            workCell = sheet.cell(row, column)
            workCell.value = templateCell.value
            workCell.fill = copy(templateCell.fill)
            workCell.font = copy(templateCell.font)
            workCell.alignment = copy(templateCell.alignment)
            workCell.border = copy(templateCell.border)
            workCell.number_format = copy(templateCell.number_format)
    sheetRow = 32
    sheetColumn = 3
    for row in range(sheetRow, maxRow + 1):
        for column in range(sheetColumn, maxColumn + 1):
            workCell = sheet.cell(row, column)
            value = "="
            for sObject in objects:
                cellCoordinate = workCell.coordinate
                value += "'" + sObject + "'!" + cellCoordinate + "+"
            workCell.value = value[:len(value) - 1]

def fillC2(sheet, objects, templateSheet):
    maxColumn = templateSheet.max_column
    maxRow = templateSheet.max_row
    for row in range(1, maxRow + 1):
        for column in range(1, maxColumn):
            templateCell = templateSheet.cell(row, column)
            workCell = sheet.cell(row, column);
            workCell.value = templateCell.value
            workCell.fill = copy(templateCell.fill)
            workCell.font = copy(templateCell.font)
            workCell.alignment = copy(templateCell.alignment)
            workCell.border = copy(templateCell.border)
            workCell.number_format = copy(templateCell.number_format)
    sheetRow = 32
    sheetColumn = 3
    for row in range(sheetRow, 81):
        for column in range(sheetColumn, maxColumn):
            workCell = sheet.cell(row, column)
            value = "="
            for sObject in objects:
                if sObject == "Чумаченко, 13В":
                    cellCoordinate = workCell.coordinate[:1] + "168"
                else:
                    cellCoordinate = workCell.coordinate
                value += "'" + sObject + "'!" + cellCoordinate + "+"
            workCell.value = value[:len(value) - 1]

def fillCKNS(sheet, objects, templateSheet):
    maxColumn = templateSheet.max_column
    maxRow = templateSheet.max_row
    for row in range(1, maxRow + 1):
        for column in range(1, maxColumn):
            templateCell = templateSheet.cell(row, column)
            workCell = sheet.cell(row, column)
            workCell.value = templateCell.value
            workCell.fill = copy(templateCell.fill)
            workCell.font = copy(templateCell.font)
            workCell.alignment = copy(templateCell.alignment)
            workCell.border = copy(templateCell.border)
            workCell.number_format = copy(templateCell.number_format)
    sheetRow = 32
    sheetColumn = 3
    for row in range(sheetRow, 81):
        for column in range(sheetColumn, maxColumn):
            workCell = sheet.cell(row, column)
            value = "="
            for sObject in objects:
                if sObject == "Чумаченко, 13В":
                    cellCoordinate = workCell.coordinate[:1] + "236"
                else:
                    cellCoordinate = workCell.coordinate
                value += "'" + sObject + "'!" + cellCoordinate + "+"
            workCell.value = value[:len(value) - 1]

def fillC4(sheet, objects, templateSheet):
    maxColumn = templateSheet.max_column
    maxRow = templateSheet.max_row
    for row in range(1, maxRow + 1):
        for column in range(1, maxColumn + 1):
            templateCell = templateSheet.cell(row, column)
            workCell = sheet.cell(row, column)
            if row > 31 and column > 2:
                value = "="
                for sObject in objects:
                    cellCoordinate = workCell.coordinate
                    value += "'" + sObject + "'!" + cellCoordinate + "+"
                workCell.value = value[:len(value) - 1]
            else:
                workCell.value = templateCell.value
            workCell.fill = copy(templateCell.fill)
            workCell.font = copy(templateCell.font)
            workCell.alignment = copy(templateCell.alignment)
            workCell.border = copy(templateCell.border)
            workCell.number_format = copy(templateCell.number_format)

def export_to_excel(connection, filePath, path, month, year):
    block1Length = 106

    wb = openpyxl.Workbook()
    sheet = wb.active

    connection.set_client_encoding("utf-8")
    queryString = "select idnode, nodename from t_conftree where typedenom = 'OBJECT' and nodename not like '%с/п%' and nodename not like 'тест'"
    objectsCursor = connection.cursor()
    objectsCursor.execute(queryString)
    objects = objectsCursor.fetchall()
    objectsCursor.close()
    fontRed = Font(name = 'Arial', size = 11, bold = True, italic = False, vertAlign = None, underline = 'none', strike = False, color = 'FFFF0000')
    fontBold = Font(name = 'Arial', size = 11, bold = True, italic = False, vertAlign = None, underline = 'none', strike = False, color = 'FF000000')
    greenFill = PatternFill(fill_type = "solid", start_color = "00DAFFFF", end_color = "00DAFFFF")
    
    for row in enumerate(objects, start = 2):
        objectId = row[1][0]
        objectName = row[1][1]
        print(objectName)
        sheet.title = objectName
        queryString = "select objref, nodename from t_conftree where typedenom = 'FIDER' and parentref = " + str(objectId)
        feedersCursor = connection.cursor()
        feedersCursor.execute(queryString)
        feeders = feedersCursor.fetchall()
        feedersCursor.close()
        queryString = "select idnode, nodename from t_conftree where typedenom = 'OBJECT' and nodename like '" + objectName + " с/п'"
        spCursor = connection.cursor()
        spCursor.execute(queryString)
        sp = spCursor.fetchall()
        spCursor.close()
        startDate = datetime.strptime(year + "-" + month + "-01 00:00:00", "%Y-%m-%d %H:%M:%S")
        numDays = monthrange(startDate.year, startDate.month)
        feederIdsString = ""
        subFeederIdsString = ""
        tableColumn = 1
        spTableColumn = 35
        startSelfTableRow = 1
        startSumTableRow = 54
        workCell = sheet.cell(startSelfTableRow, tableColumn)
        workCell.font = fontRed
        workCell.value = objectName + " СОБСТВЕННОЕ"
        startSelfTableRow += 2
        workCell = sheet.cell(startSumTableRow, tableColumn)
        workCell.font = fontRed
        workCell.value = "СВОДНАЯ ПО " + objectName
        workCell = sheet.cell(startSumTableRow, spTableColumn)
        workCell.font = fontRed
        workCell.value = "СВОДНАЯ ПО " + objectName + " СУБПОТРЕБИТЕЛИ"
        startSumTableRow += 1
        startDate = datetime.strptime(year + "-" + month + "-01 00:00:00", "%Y-%m-%d %H:%M:%S")
        endDate = datetime.strptime(year + "-" + month + "-" + str(numDays[1]) + " 23:59:00", "%Y-%m-%d %H:%M:%S")
        feederIdsString = fillFeeders(connection, sheet, feeders, block1Length, 1, startDate, endDate, fontBold, greenFill)
        startDate = datetime.strptime(year + "-" + month + "-01 00:00:00", "%Y-%m-%d %H:%M:%S")
        endDate = datetime.strptime(year + "-" + month + "-" + str(numDays[1]) + " 23:59:00", "%Y-%m-%d %H:%M:%S")
        for spRow in enumerate(sp, start = 2):
            subObjectId = spRow[1][0]
            queryString = "select objref, nodename from t_conftree where typedenom = 'FIDER' and parentref = " + str(subObjectId)
            subFeedersCursor = connection.cursor()
            subFeedersCursor.execute(queryString)
            subFeeders = subFeedersCursor.fetchall()
            subFeedersCursor.close()
            subFeederIdsString = fillFeeders(connection, sheet, subFeeders, block1Length, 35, startDate, endDate, fontBold, greenFill)
        pDay = 0
        selfRowSums = a.array("d", [])
        rowSums = a.array("d", [])
        spRowSums = a.array("d", [])
        index = 0
        endDate = datetime.strptime(year + "-" + month + "-01 00:00:00", "%Y-%m-%d %H:%M:%S")
        sumTableRow = startSumTableRow + 1
        while startDate.day <= numDays[1]:
            offset = 0
            if pDay != startDate.day:
                sumText = "Всего"
                if tableColumn == 1:
                    selfCell = sheet.cell(startSelfTableRow, tableColumn)
                    sumCell = sheet.cell(startSumTableRow, tableColumn)
                    spSumCell = sheet.cell(startSumTableRow, spTableColumn)
                    selfCell.fill = greenFill
                    sumCell.fill = greenFill
                    spSumCell.fill = greenFill
                    timeLabelText = "Время"
                    selfCell.value = timeLabelText
                    sumCell.value = timeLabelText
                    spSumCell.value = timeLabelText
                    tableColumn += 1
                    spTableColumn += 1
                if pDay != 0:
                    workCell = sheet.cell(sumTableRow, tableColumn)
                    workCell.value = daySum
                    workCell = sheet.cell(sumTableRow, spTableColumn)
                    workCell.value = spDaySum
                    workCell = sheet.cell(selfTableRow, tableColumn)
                    workCell.value = daySum - spDaySum
                    index = 0
                    if tableColumn == 2:
                        workCell = sheet.cell(sumTableRow, tableColumn - 1)
                        workCell.value = sumText
                        workCell = sheet.cell(selfTableRow, tableColumn - 1)
                        workCell.value = sumText
                    tableColumn += 1
                    spTableColumn += 1
                    workCell = sheet.cell(sumTableRow, spTableColumn)
                    workCell.value = sumText
                selfCell = sheet.cell(startSelfTableRow, tableColumn)
                sumCell = sheet.cell(startSumTableRow, tableColumn)
                spSumCell = sheet.cell(startSumTableRow, spTableColumn)
                selfCell.fill = greenFill
                sumCell.fill = greenFill
                spSumCell.fill = greenFill
                selfCell.value = startDate.strftime("%d") + "." + startDate.strftime("%m") + "." + startDate.strftime("%Y")[2:]
                sumCell.value = startDate.strftime("%d") + "." + startDate.strftime("%m") + "." + startDate.strftime("%Y")[2:]
                spSumCell.value = startDate.strftime("%d") + "." + startDate.strftime("%m") + "." + startDate.strftime("%Y")[2:]
                if startDate.day == numDays[1]:
                    selfCell = sheet.cell(startSelfTableRow, tableColumn + 1)
                    sumCell = sheet.cell(startSumTableRow, tableColumn + 1)
                    spSumCell = sheet.cell(startSumTableRow, spTableColumn + 1)
                    selfCell.fill = greenFill
                    sumCell.fill = greenFill
                    spSumCell.fill = greenFill
                    selfCell.value = sumText
                    sumCell.value = sumText
                    spSumCell.value = sumText
                    sumCell = sheet.cell(sumTableRow + 1, tableColumn + 1)
                if pDay == 0:
                    pDay = startDate.day
                sumTableRow = startSumTableRow + 1
                selfTableRow = startSelfTableRow + 1
                daySum = 0
                spDaySum = 0
            queryString = "select dt, sum(real_val) from d_gr30 where dt >= '" + startDate.strftime("%Y-%m-%d %H:%M:%S") + "' and dt <= '" + endDate.strftime("%Y-%m-%d %H:%M:%S") + "' and idfiders in (" + feederIdsString[0:-1] +") and dr = 1 GROUP BY dt"
            spGr30 = a.array("i", [])
            if len(subFeederIdsString) > 0:
                spQueryString = "select dt, sum(real_val) from d_gr30 where dt >= '" + startDate.strftime("%Y-%m-%d %H:%M:%S") + "' and dt <= '" + endDate.strftime("%Y-%m-%d %H:%M:%S")
                spQueryString += "' and idfiders in (" + subFeederIdsString[0:-1] +") and dr = 1 GROUP BY dt"
                spDataCursor = connection.cursor()
                spDataCursor.execute(spQueryString)
                spGr30 = spDataCursor.fetchall()
                spDataCursor.close()
            dataCursor = connection.cursor() 
            dataCursor.execute(queryString)
            gr30 = dataCursor.fetchall()
            dataCursor.close()
            if len(gr30) > 0:
                if tableColumn == 2:
                    timeValue = (gr30[0][0] + timedelta(minutes = 30)).strftime("%H:%M")
                    timeCell = sheet.cell(sumTableRow, tableColumn - 1)
                    timeCell.value = timeValue
                    timeCell = sheet.cell(selfTableRow, tableColumn - 1)
                    timeCell.value = timeValue
                    timeCell = sheet.cell(sumTableRow, spTableColumn - 1)
                    timeCell.value = timeValue
                value = gr30[0][1]
            else:
                value = 0
            if len(spGr30) > 0:
                spValue = spGr30[0][1]
            else:
                spValue = 0
            if value == None:
                value = 0
            if spValue == None:
                spValue = 0
            workCell = sheet.cell(sumTableRow, tableColumn)
            spCell = sheet.cell(sumTableRow, spTableColumn)
            sCell = sheet.cell(selfTableRow, tableColumn)
            workCell.value = value
            spCell.value = spValue
            sCell.value = value - spValue
            daySum += value
            spDaySum += spValue
            if tableColumn == 2:
                selfRowSums.append(0)
                rowSums.append(0)
                spRowSums.append(0)
            rowSums.insert(index, rowSums[index] + value)
            spRowSums.insert(index, spRowSums[index] + spValue)
            selfRowSums.insert(index, selfRowSums[index] + (value - spValue))
            if startDate.day == numDays[1]:
                workCell = sheet.cell(sumTableRow, tableColumn + 1)
                workCell.value = rowSums[index]
                workCell = sheet.cell(sumTableRow, spTableColumn + 1)
                workCell.value = spRowSums[index]
                workCell = sheet.cell(selfTableRow, tableColumn + 1)
                workCell.value = selfRowSums[index]
            index += 1
            sumTableRow += 1
            selfTableRow += 1
            startSumTableRow = 55
            pDay = startDate.day
            if startDate.day == numDays[1] and startDate.hour == 23 and startDate.minute >= 30:
                workCell = sheet.cell(sumTableRow, tableColumn)
                workCell.value = daySum
                workCell = sheet.cell(sumTableRow, spTableColumn)
                workCell.value = spDaySum
                workCell = sheet.cell(selfTableRow, tableColumn)
                workCell.value = daySum - spDaySum
                break
            startDate = startDate + timedelta(minutes = 30)
            endDate = endDate + timedelta(minutes = 30)
        sheet = wb.create_sheet("new")
    vsObjects = ["Чумаченко, 13В", "ДВС-1", "ДВС-2", "НС Хортицкая", "НС Леваневская", "НС Шевченко", "НС Павло-Кичкас", "НС Балабино", "НС \"Соцгород\"", "НС Товарищеская"]
    fillSvod(sheet, "Водоснабжение", vsObjects, 211, greenFill)
    sheet = wb.create_sheet("new")
    voObjects = ["Чумаченко, 13В", "ЦОС-1", "ЦОС-2", "КНС-1", "КНС-2", "КНС-7", "КНС-9", "КНС-23", "КНС-24", "КНС-6", "КНС-4", "КНС-30", "КНС-29", "КНС-25", "КНС-22", "КНС-39", "КНС-3"]
    fillSvod(sheet, "Водоотведение", voObjects, 159, greenFill)
    sheet = wb.create_sheet("new")
    adminObjects = ["Чумаченко, 13В", "Промбаза", "Св. Николая, 61"]
    fillSvod(sheet, "Админ", adminObjects, 107, greenFill)
    sheet = wb.create_sheet("new")
    subObjects = ["ДВС-1", "ДВС-2", "ЦОС-1", "ЦОС-2", "НС Хортицкая", "НС Леваневская", "НС Шевченко", "КНС-1", "КНС-2", "КНС-7", "КНС-9", "КНС-23", "КНС-24", "Промбаза", "НС Павло-Кичкас",
                  "КНС-6", "НС Балабино", "Св. Николая, 61", "КНС-4", "КНС-30", "КНС-29", "КНС-25", "НС \"Соцгород\"", "КНС-22", "НС Товарищеская", "Чумаченко, 13В", "КНС-39", "КНС-3"]
    fillSubSvod(sheet, "Субпотребители", subObjects, wb.worksheets[1])
    startDate = datetime.strptime(year + "-" + month + "-01 00:00:00", "%Y-%m-%d %H:%M:%S")
    endDate = datetime.strptime(year + "-" + month + "-" + str(numDays[1]) + " 23:59:00", "%Y-%m-%d %H:%M:%S")
    fillNoASKUE(connection, wb, startDate, endDate, fontRed, greenFill)
    svodObjects = ["Водоснабжение", "Водоотведение", "Админ", "Суб без АСКУЭ договор 10", "Суб без АСКУЭ договор 15", "Суб без АСКУЭ договор 2"]
    fillSvodSheet(connection, wb, startDate, endDate, fontRed, greenFill)
    startDate = datetime.strptime(year + "-" + month + "-01 00:00:00", "%Y-%m-%d %H:%M:%S")
    wb.save(filePath + "\\" + startDate.strftime("%B") + "." + str(startDate.year) +".xlsx")
main()
